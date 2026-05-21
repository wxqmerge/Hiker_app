import json
import re
from pathlib import Path
from openpyxl import Workbook


def month_to_quarters(months):
    quarters = {}
    q1_months = {3, 4, 5}
    q2_months = {6, 7, 8}
    q3_months = {9, 10, 11}
    q4_months = {12, 1, 2}

    if any(m in q1_months for m in months):
        quarters['Q1'] = '1'
    if any(m in q2_months for m in months):
        quarters['Q2'] = '1'
    if any(m in q3_months for m in months):
        quarters['Q3'] = '1'
    if any(m in q4_months for m in months):
        quarters['Q4'] = '1'

    return quarters


def main():
    data_path = Path('export_for_excel.json')

    if not data_path.exists():
        print("Error: export_for_excel.json not found. Run 'Export for Excel' from the app first.")
        return

    with open(data_path, encoding='utf-8') as f:
        data = json.load(f)

    trails = data.get('trails', [])
    details_data = data.get('trail_details', {})

    wb = Workbook()

    ws_index = wb.active
    ws_index.title = 'Index'

    COL_FULL_NAME = 0
    COL_DISTANCE = 1
    COL_DIST_EXT = 2
    COL_ELEV_START = 3
    COL_ELEV_MAX = 4
    COL_RANGE = 6
    COL_Q1 = 7
    COL_Q2 = 8
    COL_Q3 = 11
    COL_Q4 = 15
    COL_DIFFICULTY = 17
    COL_SHORT_NAME = 18

    ws_index.cell(row=1, column=COL_FULL_NAME + 1, value='Full Name')
    ws_index.cell(row=1, column=COL_DISTANCE + 1, value='Distance')
    ws_index.cell(row=1, column=COL_DIST_EXT + 1, value='Distance Extended')
    ws_index.cell(row=1, column=COL_ELEV_START + 1, value='Elevation Start')
    ws_index.cell(row=1, column=COL_ELEV_MAX + 1, value='Elevation Max')
    ws_index.cell(row=1, column=COL_RANGE + 1, value='Range')
    ws_index.cell(row=1, column=COL_DIFFICULTY + 1, value='Difficulty')
    ws_index.cell(row=1, column=COL_SHORT_NAME + 1, value='Short Name')

    for idx, trail in enumerate(trails, start=2):
        row = idx

        seasonal = trail.get('seasonal', {})
        available_months = seasonal.get('availableMonths', [])
        quarters = month_to_quarters(available_months)

        ws_index.cell(row=row, column=COL_FULL_NAME + 1, value=trail.get('fullName', trail.get('name', '')))
        ws_index.cell(row=row, column=COL_DISTANCE + 1, value=trail.get('distance'))
        ws_index.cell(row=row, column=COL_DIST_EXT + 1, value=trail.get('distanceExtended'))
        ws_index.cell(row=row, column=COL_ELEV_START + 1, value=trail.get('elevationStart'))
        ws_index.cell(row=row, column=COL_ELEV_MAX + 1, value=trail.get('elevationMax'))
        ws_index.cell(row=row, column=COL_RANGE + 1, value=trail.get('range'))
        ws_index.cell(row=row, column=COL_Q1 + 1, value=quarters.get('Q1'))
        ws_index.cell(row=row, column=COL_Q2 + 1, value=quarters.get('Q2'))
        ws_index.cell(row=row, column=COL_Q3 + 1, value=quarters.get('Q3'))
        ws_index.cell(row=row, column=COL_Q4 + 1, value=quarters.get('Q4'))
        ws_index.cell(row=row, column=COL_DIFFICULTY + 1, value=trail.get('difficulty', 'Unknown'))
        ws_index.cell(row=row, column=COL_SHORT_NAME + 1, value=trail.get('name', ''))

        trail_id = trail.get('id', '')
        sheet_name = trail.get('name', f'sheet{idx}')
        sheet_name = re.sub(r'[/\\?*:\[\]]', '', sheet_name)[:31]
        if not sheet_name:
            sheet_name = f'trail{idx}'

        existing_names = [ws.title for ws in wb.worksheets]
        if sheet_name in existing_names:
            counter = 2
            while f'{sheet_name}_{counter}' in existing_names:
                counter += 1
            sheet_name = f'{sheet_name}_{counter}'[:31]

        ws_detail = wb.create_sheet(title=sheet_name)

        ws_detail.cell(row=1, column=1, value=sheet_name)

        detail = details_data.get(trail_id, {})

        parking = trail.get('parking') or detail.get('parking')
        if parking:
            ws_detail.cell(row=4, column=2, value=parking)

        range_val = trail.get('range') or detail.get('range')
        if range_val:
            ws_detail.cell(row=5, column=7, value=range_val)

        description = detail.get('fullDescription', '')
        if description:
            words = description.split()
            lines = []
            current_line = []
            current_length = 0
            for word in words:
                if current_length + len(word) + 1 > 80:
                    lines.append(' '.join(current_line))
                    current_line = [word]
                    current_length = len(word)
                else:
                    current_line.append(word)
                    current_length += len(word) + 1
            if current_line:
                lines.append(' '.join(current_line))

            for i, line in enumerate(lines[:12]):
                ws_detail.cell(row=7 + i, column=1, value=line)

        pros = detail.get('pros')
        if pros:
            ws_detail.cell(row=14, column=2, value=pros)

        others = detail.get('others')
        if others:
            ws_detail.cell(row=17, column=2, value=others)

        leaders = detail.get('leaders', [])
        if leaders:
            ws_detail.cell(row=20, column=2, value=', '.join(leaders))

    output_path = Path('hiker_export.xlsx')
    wb.save(output_path)
    print(f"Exported {len(trails)} trails to {output_path}")


if __name__ == '__main__':
    main()
